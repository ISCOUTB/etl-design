import csv
import openpyxl
from io import BytesIO

import services.dtypes as dtypes
from typing import Any, List, Dict


def open_file_from_bytes(file_bytes: bytes, **kwargs: Any) -> openpyxl.Workbook:
    """
    Open an Excel file from bytes.

    Args:
        file_bytes (bytes): The bytes of the Excel file.
        **kwargs: Additional keyword arguments to pass to openpyxl.load_workbook.

    Returns:
        openpyxl.Workbook: The loaded workbook object.
    """
    excel_file = BytesIO(file_bytes)
    return openpyxl.load_workbook(excel_file, data_only=False, **kwargs)


def extract_formulas(
    workbook: openpyxl.Workbook,
) -> Dict[str, Dict[str, List[dtypes.CellData]]]:
    """
    Extract formulas and cell data from an Excel workbook.

    Args:
        workbook (openpyxl.Workbook): The workbook object containing the Excel data.

    Returns:
        List[dtypes.CellData]: A list of dictionaries containing cell data, including sheet name,
        cell coordinate, value, data type, and whether the cell contains a formula.
    """
    sheets: Dict[str, Dict[str, List[dtypes.CellData]]] = {}

    for sheet in workbook.worksheets:
        sheets[sheet.title] = {}
        for column in sheet.columns:
            column_letter = column[0].column_letter
            result: List[dtypes.CellData] = []
            for cell in column:
                # print(cell.coordinate, cell.column, cell.row, cell.col_idx, cell.column_letter)
                cell_data: dtypes.CellData = {
                    "cell": cell.coordinate,
                    "value": cell.value,
                    "data_type": cell.data_type,
                    "is_formula": cell.data_type
                    == "f",  # isinstance(cell.value, str) and cell.value.startswith("=")
                }
                result.append(cell_data)
            sheets[sheet.title][column_letter] = result

    return sheets


def convert_csv_to_excel(file_bytes: bytes) -> openpyxl.Workbook:
    """
    Convert CSV file bytes to Excel workbook.

    Args:
        file_bytes (bytes): The bytes of the CSV file.

    Returns:
        openpyxl.Workbook: The Excel workbook object.
    """
    csv_text = file_bytes.decode("utf-8")
    csv_reader = csv.reader(csv_text.splitlines())
    csv_data = list(csv_reader)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"

    for row_idx, row_data in enumerate(csv_data, start=1):
        for col_idx, cell_value in enumerate(row_data, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=cell_value)

    return workbook
