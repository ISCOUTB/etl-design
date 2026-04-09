from io import BytesIO

import openpyxl

from src.services.utils import (
    convert_csv_to_excel,
    extract_cell_data,
    open_file_from_bytes,
)


def test_open_file_from_bytes_calls_openpyxl(monkeypatch):
    called = {}

    def _fake_load_workbook(file_obj, data_only=False, **kwargs):
        called["is_bytesio"] = isinstance(file_obj, BytesIO)
        called["data_only"] = data_only
        called["kwargs"] = kwargs
        return "workbook"

    monkeypatch.setattr(
        "src.services.utils.openpyxl.load_workbook", _fake_load_workbook
    )

    result = open_file_from_bytes(b"fake-bytes", read_only=True)

    assert result == "workbook"
    assert called["is_bytesio"] is True
    assert called["data_only"] is False
    assert called["kwargs"]["read_only"] is True


def test_convert_csv_to_excel_creates_sheet_and_cells():
    csv_bytes = b"id,name\n1,Alice\n2,Bob\n"

    workbook = convert_csv_to_excel(csv_bytes)

    sheet = workbook.active
    assert sheet is not None
    assert sheet.title == "Sheet1"
    assert sheet["A1"].value == "id"
    assert sheet["B1"].value == "name"
    assert sheet["A2"].value == "1"
    assert sheet["B3"].value == "Bob"


def test_extract_cell_data_detects_formulas_and_limit():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Sheet1"

    # Header + 2 rows
    sheet["A1"] = "amount"
    sheet["A2"] = 10
    sheet["A3"] = "=A2*2"

    result = extract_cell_data(workbook, limit=2)

    assert "Sheet1" in result
    assert "A" in result["Sheet1"]
    # limit=2 => header + first row only
    assert len(result["Sheet1"]["A"]) == 2
    assert result["Sheet1"]["A"][0]["cell"] == "A1"
    assert result["Sheet1"]["A"][0]["is_formula"] is False
    assert result["Sheet1"]["A"][1]["value"] == 10
