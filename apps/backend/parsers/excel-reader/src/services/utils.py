import csv
from io import BytesIO
from typing import Any, Dict, List, Optional

import openpyxl

from src import schemas
from src.utils.monitor_performance import monitor_performance


@monitor_performance("open_file_from_bytes")
def open_file_from_bytes(file_bytes: bytes, **kwargs: Any) -> openpyxl.Workbook:
    """
    Open an Excel file from bytes.

    Args:
        file_bytes (bytes): The bytes of the Excel file.
        **kwargs: Additional keyword arguments to pass to openpyxl.load_workbook.

    Returns:
        openpyxl.Workbook: The loaded workbook object.
    """
    excel_file = BytesIO(file_bytes)
    return openpyxl.load_workbook(excel_file, data_only=False, **kwargs)


def extract_cell_data(
    workbook: openpyxl.Workbook, limit: Optional[int] = None
) -> Dict[str, Dict[str, List[schemas.CellData]]]:
    """
    Extract formulas and cell data from an Excel workbook.

    Args:
        workbook (openpyxl.Workbook): The workbook object containing the Excel data.
        limit (int): The maximum number of cells to extract per column.

    Returns:
        List[schemas.CellData]: A list of dictionaries containing cell data, including sheet name,
        cell coordinate, value, data type, and whether the cell contains a formula.
    """
    sheets: Dict[str, Dict[str, List[schemas.CellData]]] = {}

    for sheet in workbook.worksheets:
        if sheet.sheet_state != "visible":
            continue
        sheets[sheet.title] = {}
        if limit is None or limit <= 0:
            limit = sheet.max_row

        max_rows = sheet.max_row if limit <= 0 else min(limit, sheet.max_row)
        for column in sheet.columns:
            if column[0].value is None:
                continue

            column_letter = column[0].column_letter
            result: List[schemas.CellData] = []
            for cell, _ in zip(column, range(max_rows)):
                cell_data: schemas.CellData = {
                    "cell": cell.coordinate,
                    "value": cell.value,
                    "data_type": cell.data_type,
                    # isinstance(cell.value, str) and cell.value.startswith("=")
                    "is_formula": cell.data_type == "f",
                }
                result.append(cell_data)
            sheets[sheet.title][column_letter] = result

    # {sheet_name: {column_letter: [cell_data, ...], ...}, ...}
    return sheets


def convert_csv_to_excel(file_bytes: bytes) -> openpyxl.Workbook:
    """
    Convert CSV file bytes to Excel workbook.

    Args:
        file_bytes (bytes): The bytes of the CSV file.

    Returns:
        openpyxl.Workbook: The Excel workbook object.
    """
    csv_text = file_bytes.decode("utf-8")
    csv_reader = csv.reader(csv_text.splitlines())
    csv_data = list(csv_reader)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert sheet is not None

    sheet.title = "Sheet1"
    for row_idx, row_data in enumerate(csv_data, start=1):
        for col_idx, cell_value in enumerate(row_data, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=cell_value)

    return workbook
